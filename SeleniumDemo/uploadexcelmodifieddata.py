import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions

def update_excel_data(filepath,searchTerm,columnName,new_value):

    book = openpyxl.load_workbook(filepath)
    sheet =book.active
    Dict={}

    for i in range(1,sheet.max_column+1):
        if sheet.cell(1,column=i).value == columnName:
            Dict["col"] = i

    for i in range(1,sheet.max_row+1):
        for j in range(1,sheet.max_column+1):
            if sheet.cell(row=i,column=j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_value
    book.save(file_path)

file_path = "C:\\Users\\abina\\Downloads\\download.xlsx"
fruit_name = "Apple"
newValue = "990"
service_obj = Service("C:/Users/abina/Downloads/chromedriver-win64 (1)/chromedriver-win64/chromedriver.exe") #C:\Users\abina\Downloads\chromedriver-win64 (1)\chromedriver-win64
driver = webdriver.Chrome(service=service_obj)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.implicitly_wait(5)


# driver.find_element(By.ID,"downloadButton").click()

#edit the excel with updated value
update_excel_data(file_path,fruit_name,"price",newValue)

#upload
file_input = driver.find_element(By.CSS_SELECTOR,"input[type='file']")
file_input.send_keys(file_path)


WebDriverWait(driver,5).until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[text()='Updated Excel Data Successfully.']")))#css selector =".Toastify__toast-body div:nth-child(2)"
upload_text=driver.find_element(By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)").text
print(upload_text)
price_column=driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,f"//div[text()='{fruit_name}']/parent::div/parent::div/div[@id='cell-{price_column}-undefined']/div").text
print("The Actual price of the apple is :",actual_price)

assert  actual_price == newValue
time.sleep(2)
driver.quit()