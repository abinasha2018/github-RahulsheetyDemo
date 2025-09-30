import openpyxl
book = openpyxl.load_workbook("C:\\Users\\abina\\Desktop\\PythonDemo.xlsx")
sheet =book.active


cell = sheet.cell(1,2)
print(cell.value)

#write something to excel
sheet.cell(2,2).value = "Rahul"

print(sheet.cell(2,2).value)

#No of rows of complete sheet
print(sheet.max_row)

#No of column present in the sheet
print(sheet.max_column)

#Other way of print the cell value
print(sheet['A5'].value)

print("----------------------------------------------------")


# for i in range(1,sheet.max_row+1):#to get rows
#     if sheet.cell(row=i,column=1).value=="Testcase2" :
#         for j in range(1,sheet.max_column+1): #to get columns
#             print(sheet.cell(row=i,column=j).value ,end="  ")
#         print()
Dict={}
for i in range(1,sheet.max_row+1):#to get rows
    if sheet.cell(row=i,column=1).value=="Testcase2" :
        for j in range(1,sheet.max_column+1): #to get columns
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(Dict)








